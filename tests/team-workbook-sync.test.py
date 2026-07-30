import importlib.util
import json
import shutil
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / 'data' / '生康足球队数据源.xlsx'
SITE_DATA_PATH = ROOT / 'js' / 'team-data.js'
NEWCOMER_DATA_PATH = ROOT / 'js' / 'newcomer-data.js'
SYNC_SPEC = importlib.util.spec_from_file_location('sync_team_workbook', ROOT / 'tools' / 'sync_team_workbook.py')
SYNC_MODULE = importlib.util.module_from_spec(SYNC_SPEC)
SYNC_SPEC.loader.exec_module(SYNC_MODULE)
newcomer_payload = SYNC_MODULE.newcomer_payload


def load_window_data(path, global_name):
    script = """
const fs = require('fs');
const vm = require('vm');
const context = { window: {} };
vm.runInNewContext(fs.readFileSync(process.argv[1], 'utf8'), context);
process.stdout.write(JSON.stringify(context.window[process.argv[2]]));
"""
    result = subprocess.run(
        ['node', '-e', script, str(path), global_name],
        cwd=ROOT,
        check=True,
        capture_output=True,
    )
    return json.loads(result.stdout.decode('utf-8'))


def load_site_data(path=SITE_DATA_PATH):
    return load_window_data(path, 'PONYTAIL_DATA')


def load_newcomer_data(path=NEWCOMER_DATA_PATH):
    return load_window_data(path, 'NEWCOMER_DATA')


class TeamWorkbookSyncTest(unittest.TestCase):
    def test_workbook_has_isolated_newcomer_sheet(self):
        workbook = load_workbook(WORKBOOK_PATH, data_only=False)
        self.assertIn('新生展示', workbook.sheetnames)
        headers = [cell.value for cell in workbook['新生展示'][1]]
        self.assertEqual(headers, [
            '排序', '姓名', '年级', '号码', '主位置', '可胜任位置',
            '惯用脚', '踢球风格', '自我介绍', '照片文件名', '照片焦点', '展示状态',
        ])

    def test_sync_emits_nonempty_newcomers_without_changing_official_roster(self):
        official_source = SITE_DATA_PATH.read_bytes()
        official_data = load_site_data()
        with tempfile.TemporaryDirectory() as directory:
            temporary = Path(directory)
            source = temporary / 'source.xlsx'
            target = temporary / 'synced.xlsx'
            site_output = temporary / 'team-data.js'
            newcomer_output = temporary / 'newcomer-data.js'
            shutil.copy2(WORKBOOK_PATH, source)

            workbook = load_workbook(source)
            workbook['新生展示'].append([
                1, 'Fixture Newcomer', '2026级', 24, 'MF', 'AM',
                '右脚', '组织', 'Fixture intro', 'fixture.webp', '50% 25%', '展示',
            ])
            workbook.save(source)

            SYNC_MODULE.sync_workbook(
                source, target, SITE_DATA_PATH, site_output, newcomer_output,
            )

            newcomer_data = load_newcomer_data(newcomer_output)
            synced_site_data = load_site_data(site_output)
            self.assertEqual(newcomer_data['season'], '2026')
            self.assertEqual(
                [item['name'] for item in newcomer_data['newcomers']],
                ['Fixture Newcomer'],
            )
            self.assertEqual(synced_site_data['players'], official_data['players'])
            self.assertEqual(
                synced_site_data['startingLineup'],
                official_data['startingLineup'],
            )

        self.assertEqual(SITE_DATA_PATH.read_bytes(), official_source)

    def test_empty_newcomer_payload_has_required_shape(self):
        payload = load_newcomer_data()
        self.assertEqual(payload, {'season': '', 'newcomers': []})
        self.assertEqual(set(payload), {'season', 'newcomers'})

    def test_newcomer_payload_filters_sorts_and_defaults_fields(self):
        payload = newcomer_payload([
            {
                '排序': 2, '姓名': 'Second', '年级': '2026级', '号码': 9,
                '主位置': 'FW', '可胜任位置': 'LW', '惯用脚': '右脚',
                '踢球风格': '冲刺', '自我介绍': '第二位', '照片文件名': '',
                '照片焦点': '', '展示状态': '展示',
            },
            {
                '排序': 1, '姓名': 'First', '年级': '2026级', '号码': 7,
                '主位置': 'MF', '可胜任位置': 'AM', '惯用脚': '左脚',
                '踢球风格': '组织', '自我介绍': '第一位', '照片文件名': 'first.webp',
                '照片焦点': '50% 25%', '展示状态': '展示',
            },
            {
                '排序': 3, '姓名': 'Hidden', '年级': '2026级', '号码': 10,
                '主位置': 'DF', '可胜任位置': 'CB', '惯用脚': '右脚',
                '踢球风格': '防守', '自我介绍': '不展示', '照片文件名': 'hidden.webp',
                '照片焦点': '50% 50%', '展示状态': '隐藏',
            },
        ])
        self.assertEqual(payload, {
            'season': '2026',
            'newcomers': [
                {
                    'order': 1, 'name': 'First', 'grade': '2026级', 'number': 7,
                    'pos': 'MF', 'role': 'AM', 'preferredFoot': '左脚',
                    'style': '组织', 'intro': '第一位', 'photo': 'assets/players/first.webp',
                    'photoPosition': '50% 25%',
                },
                {
                    'order': 2, 'name': 'Second', 'grade': '2026级', 'number': 9,
                    'pos': 'FW', 'role': 'LW', 'preferredFoot': '右脚',
                    'style': '冲刺', 'intro': '第二位', 'photo': '', 'photoPosition': '50% 50%',
                },
            ],
        })

    def test_sync_keeps_official_formula_caches_usable(self):
        tracked_files = [WORKBOOK_PATH, SITE_DATA_PATH, NEWCOMER_DATA_PATH]
        before = {path: path.read_bytes() for path in tracked_files}
        summary_cells = ['B4', 'B5', 'B6', 'B7', 'B8']
        summary_values = [17, 110, 13, 9, 0]
        with tempfile.TemporaryDirectory() as directory:
            temporary = Path(directory)
            source = temporary / 'source.xlsx'
            target = temporary / 'synced.xlsx'
            site_output = temporary / 'team-data.js'
            newcomer_output = temporary / 'newcomer-data.js'
            shutil.copy2(WORKBOOK_PATH, source)
            source_formulas = load_workbook(source, data_only=False)['统计汇总']
            expected_formulas = [source_formulas[cell].value for cell in summary_cells]
            SYNC_MODULE.write_formula_caches(
                source,
                '统计汇总',
                dict(zip(summary_cells, summary_values)),
            )
            source_cached = load_workbook(source, data_only=True)['统计汇总']
            self.assertEqual(
                [source_cached[cell].value for cell in summary_cells],
                summary_values,
            )

            SYNC_MODULE.sync_workbook(
                source, target, SITE_DATA_PATH, site_output, newcomer_output,
            )

            formula_workbook = load_workbook(target, data_only=False)
            cached_workbook = load_workbook(target, data_only=True)
            worksheet = cached_workbook['球员数据']
            headers = [cell.value for cell in worksheet[1]]
            rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
            site_players = {player['name']: player for player in load_site_data()['players']}

            self.assertEqual(
                {row['姓名']: row['评分'] for row in rows},
                {name: player['rating'] for name, player in site_players.items()},
            )
            self.assertEqual(
                {row['姓名']: row['代表数据'] for row in rows},
                {name: player['memory'] for name, player in site_players.items()},
            )
            self.assertEqual(
                [formula_workbook['统计汇总'][cell].value for cell in summary_cells],
                expected_formulas,
            )
            self.assertEqual(
                [cached_workbook['统计汇总'][cell].value for cell in summary_cells],
                summary_values,
            )
            self.assertTrue(site_output.exists())
            self.assertTrue(newcomer_output.exists())

        self.assertEqual({path: path.read_bytes() for path in tracked_files}, before)

    def test_workbook_has_formula_driven_player_data_and_schedule_sheets(self):
        self.assertTrue(WORKBOOK_PATH.exists())
        workbook = load_workbook(WORKBOOK_PATH, data_only=False)
        self.assertIn('球员数据', workbook.sheetnames)
        self.assertIn('赛事索引', workbook.sheetnames)
        self.assertIn('赛程', workbook.sheetnames)
        headers = [cell.value for cell in workbook['球员数据'][1]]
        self.assertNotIn('扑救', headers)
        self.assertIn('低失球场', headers)
        rating_formula = str(workbook['球员数据']['M2'].value)
        self.assertTrue(rating_formula.startswith('=ROUND(MIN('))
        self.assertIn('IF(E2="DF"', rating_formula)
        self.assertTrue(str(workbook['球员数据']['O2'].value).startswith('=IF('))

    def test_player_stats_in_workbook_match_the_website(self):
        workbook = load_workbook(WORKBOOK_PATH, data_only=True)
        worksheet = workbook['球员数据']
        headers = [cell.value for cell in worksheet[1]]
        rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
        site_players = {player['name']: player for player in load_site_data()['players']}

        self.assertEqual(set(site_players), {row['姓名'] for row in rows})
        for row in rows:
            player = site_players[row['姓名']]
            self.assertEqual(player['number'], row['号码'])
            self.assertEqual(player['pos'], row['主位置'])
            self.assertEqual(player['role'], row['可胜任位置'])
            self.assertEqual(player['apps'], row['出场'])
            self.assertEqual(player['goals'], row['进球'])
            self.assertEqual(player['asts'], row['助攻'])
            self.assertEqual(player['motm'], row['MVP'])
            self.assertEqual(player['cleanSheets'], row['零封'])
            self.assertEqual(player['lowConcedeGames'], row['低失球场'])
            self.assertEqual(player['rating'], row['评分'])
            self.assertEqual(player['memory'], row['代表数据'])
            self.assertNotIn('saves', player)


if __name__ == '__main__':
    unittest.main()
