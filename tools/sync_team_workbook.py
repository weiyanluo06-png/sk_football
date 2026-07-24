"""Synchronize the football team workbook into the static site data."""

import argparse
import json
import re
import shutil
import subprocess
import tempfile
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / 'data' / '生康足球队数据源.xlsx'
PLAYER_HEADERS = ['序号', '姓名', '绰号', '号码', '主位置', '可胜任位置', '出场', '进球', '助攻', 'MVP', '扑救', '零封', '评分', '踢球风格', '代表数据']
COMPETITION_HEADERS = ['赛事 ID', '年份', '赛事名称', '显示名称', '状态', '赛事说明']
MATCH_HEADERS = ['赛季', '赛事名称', '日期', '对手', '比分', '赛果', '场地', '进球者', '赛事备注']
GREEN = '1F7A4D'
PAPER = 'FFFDF8'
LIGHT_GREEN = 'EAF3EC'
WHITE = 'FFFFFF'
LINE = Side(style='thin', color='B9D7C2')


def site_data():
    script = """
const fs = require('fs');
const vm = require('vm');
const context = { window: {} };
vm.runInNewContext(fs.readFileSync('js/team-data.js', 'utf8'), context);
process.stdout.write(JSON.stringify(context.window.PONYTAIL_DATA));
"""
    result = subprocess.run(['node', '-e', script], cwd=ROOT, check=True, capture_output=True)
    return json.loads(result.stdout.decode('utf-8'))


def rating_formula(row):
    return f'=ROUND(MIN(9.5,MAX(6,6+G{row}*0.08+H{row}*0.35+I{row}*0.25+J{row}*0.25+K{row}*0.03+L{row}*0.15)),1)'


def memory_formula(row):
    return f'=IF(H{row}>0,IF(I{row}>0,"贡献 "&H{row}&" 粒进球 / 送出 "&I{row}&" 次助攻","贡献 "&H{row}&" 粒进球"),IF(I{row}>0,"送出 "&I{row}&" 次助攻","累计出场 "&G{row}&" 场"))'


def calculate_rating(row):
    score = Decimal('6')
    score += Decimal(str(row['出场'])) * Decimal('0.08')
    score += Decimal(str(row['进球'])) * Decimal('0.35')
    score += Decimal(str(row['助攻'])) * Decimal('0.25')
    score += Decimal(str(row['MVP'])) * Decimal('0.25')
    score += Decimal(str(row['扑救'])) * Decimal('0.03')
    score += Decimal(str(row['零封'])) * Decimal('0.15')
    return float(min(Decimal('9.5'), max(Decimal('6'), score)).quantize(Decimal('0.1'), rounding=ROUND_HALF_UP))


def calculate_memory(row):
    if row['进球'] > 0 and row['助攻'] > 0:
        return f"贡献 {row['进球']} 粒进球 / 送出 {row['助攻']} 次助攻"
    if row['进球'] > 0:
        return f"贡献 {row['进球']} 粒进球"
    if row['助攻'] > 0:
        return f"送出 {row['助攻']} 次助攻"
    return f"累计出场 {row['出场']} 场"


def style_table(sheet, width_map):
    sheet.freeze_panes = 'A2'
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f'A1:{chr(64 + sheet.max_column)}{sheet.max_row}'
    for cell in sheet[1]:
        cell.fill = PatternFill('solid', fgColor=GREEN)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor=PAPER if cell.row % 2 == 0 else LIGHT_GREEN)
            cell.border = Border(bottom=LINE)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    for column, width in width_map.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[1].height = 26


def append_table(sheet, name):
    for table_name in list(sheet.tables):
        del sheet.tables[table_name]
    table = Table(displayName=name, ref=f'A1:{chr(64 + sheet.max_column)}{sheet.max_row}')
    table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium4', showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    sheet.add_table(table)


def ensure_schedule_sheets(workbook, data):
    if '赛事索引' not in workbook.sheetnames:
        sheet = workbook.create_sheet('赛事索引')
        sheet.append(COMPETITION_HEADERS)
        for item in data.get('competitions', []):
            sheet.append([item['id'], item['year'], item['name'], item['label'], item['status'], item['description']])
    if '赛程' not in workbook.sheetnames:
        sheet = workbook.create_sheet('赛程')
        sheet.append(MATCH_HEADERS)
        for item in data.get('matches', []):
            sheet.append([item['season'], item['competition'], item['date'], item['opponent'], item['score'], item['result'], item['venue'], item['scorers'], item['quote']])

    competitions = workbook['赛事索引']
    matches = workbook['赛程']
    style_table(competitions, {'A': 24, 'B': 10, 'C': 20, 'D': 20, 'E': 18, 'F': 58})
    style_table(matches, {'A': 14, 'B': 20, 'C': 14, 'D': 16, 'E': 12, 'F': 10, 'G': 22, 'H': 28, 'I': 44})
    append_table(competitions, 'CompetitionIndex')
    append_table(matches, 'MatchSchedule')


def read_rows(sheet):
    headers = [cell.value for cell in sheet[1]]
    return [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=2, values_only=True) if values[0] is not None]


def update_players(data, rows):
    existing = {player['name']: player for player in data['players']}
    updated = []
    for row in rows:
        player = existing.get(row['姓名'], {
            'id': row['序号'], 'avatarIcon': 'fa-futbol', 'reviews': [], 'bio': '待补充球员介绍。', 'photo': '',
        })
        player.update({
            'id': row['序号'], 'name': row['姓名'], 'nickname': row['绰号'], 'number': row['号码'],
            'pos': row['主位置'], 'role': row['可胜任位置'], 'apps': row['出场'], 'goals': row['进球'],
            'asts': row['助攻'], 'motm': row['MVP'], 'saves': row['扑救'], 'cleanSheets': row['零封'],
            'rating': calculate_rating(row), 'traits': [row['踢球风格']], 'memory': calculate_memory(row),
        })
        updated.append(player)
    data['players'] = updated


def update_competitions(data, rows):
    data['competitions'] = [{
        'id': row['赛事 ID'], 'year': str(row['年份']), 'name': row['赛事名称'], 'label': row['显示名称'],
        'status': row['状态'], 'description': row['赛事说明'],
    } for row in rows]


def update_matches(data, rows):
    data['matches'] = [{
        'season': row['赛季'], 'result': row['赛果'], 'score': row['比分'], 'opponent': row['对手'],
        'date': row['日期'].strftime('%Y-%m-%d') if hasattr(row['日期'], 'strftime') else str(row['日期']),
        'venue': row['场地'], 'competition': row['赛事名称'], 'scorers': row['进球者'], 'quote': row['赛事备注'],
    } for row in rows]


def write_site_data(data):
    path = ROOT / 'js' / 'team-data.js'
    source = path.read_text(encoding='utf-8')
    players_json = json.dumps(data['players'], ensure_ascii=False, indent=8)
    competitions_json = json.dumps(data['competitions'], ensure_ascii=False, indent=8)
    matches_json = json.dumps(data['matches'], ensure_ascii=False, indent=8)
    source = re.sub(r'players:\s*\[[\s\S]*?\n    \],\n\n    startingLineup:', f'players: {players_json},\n\n    startingLineup:', source, count=1)
    source = re.sub(r'competitions:\s*\[[\s\S]*?\n    \],\n\n    matches:', f'competitions: {competitions_json},\n\n    matches:', source, count=1)
    source = re.sub(r'matches:\s*\[[\s\S]*?\n    \],\n\n    honors:', f'matches: {matches_json},\n\n    honors:', source, count=1)
    path.write_text(source, encoding='utf-8')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--source', type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument('--target', type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    args.target.parent.mkdir(parents=True, exist_ok=True)
    if args.source.resolve() != args.target.resolve():
        shutil.copy2(args.source, args.target)

    data = site_data()
    workbook = load_workbook(args.target)
    players = workbook['球员数据']
    for row in range(2, players.max_row + 1):
        players.cell(row, 13).value = rating_formula(row)
        players.cell(row, 15).value = memory_formula(row)
        players.cell(row, 13).number_format = '0.0'
    style_table(players, {'A': 8, 'B': 12, 'C': 16, 'D': 8, 'E': 10, 'F': 17, 'G': 9, 'H': 9, 'I': 9, 'J': 9, 'K': 9, 'L': 9, 'M': 12, 'N': 20, 'O': 28})
    append_table(players, 'PlayerStats')
    ensure_schedule_sheets(workbook, data)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(args.target)

    player_rows = read_rows(players)
    update_players(data, player_rows)
    update_competitions(data, read_rows(workbook['赛事索引']))
    update_matches(data, read_rows(workbook['赛程']))
    write_site_data(data)


if __name__ == '__main__':
    main()
